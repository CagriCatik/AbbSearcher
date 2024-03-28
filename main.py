import tkinter as tk
from tkinter import ttk
from tkinter import scrolledtext
from tkinter import messagebox
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from openpyxl import load_workbook

# Funktion zum Lesen der Datei und Erstellen eines Wörterbuchs
def datei_in_woerterbuch_laden(dateipfad):
    abkuerzungen_beschreibungen = {}

    try:
        # Versuchen, die Datei als XLSX zu öffnen
        workbook = load_workbook(filename=dateipfad)

        for blatt_name in workbook.sheetnames:
            sheet = workbook[blatt_name]

            for row in sheet.iter_rows(min_row=2, values_only=True):
                if len(row) >= 3:  # Annahme: Es gibt mindestens 3 Spalten (Abbreviation, Description, Detailed Description)
                    abkuerzung = row[0].lower()
                    beschreibung = row[1]
                    detailed_description = row[2]

                    if abkuerzung in abkuerzungen_beschreibungen:
                        abkuerzungen_beschreibungen[abkuerzung].append((beschreibung, detailed_description))
                    else:
                        abkuerzungen_beschreibungen[abkuerzung] = [(beschreibung, detailed_description)]

    except Exception as e:
        print(f"Fehler beim Lesen der Datei: {e}")

    return abkuerzungen_beschreibungen

# Funktion für die Suche
def suche_beschreibung(event=None):
    abkuerzung = eingabe_abkuerzung_var.get().lower()
    beschreibungen_text.config(state=tk.NORMAL)
    beschreibungen_text.delete(1.0, tk.END)

    if abkuerzung in abkuerzungen_beschreibungen:
        beschreibungen = abkuerzungen_beschreibungen[abkuerzung]
        for idx, (beschreibung, _) in enumerate(beschreibungen, start=1):
            beschreibungen_text.insert(tk.END, f"{idx}. {beschreibung}\n")

        # Zeige die detaillierte Beschreibung automatisch im neuen Fenster an
        if beschreibungen and beschreibungen[0][1]:  # Prüfe, ob "detailed description" vorhanden ist
            zeige_detaillierte_beschreibung(beschreibungen[0][1])

    else:
        beschreibungen_text.insert(tk.END, "Abkürzung nicht gefunden")

    beschreibungen_text.config(state=tk.DISABLED)


# Funktion, um detaillierte Beschreibung anzuzeigen
def zeige_detaillierte_beschreibung(detailed_description):
    # Neues Fenster für die detaillierte Beschreibung
    detailed_window = tk.Toplevel(root)
    detailed_window.title("Detailed Description")
    
    # Textfeld für die detaillierte Beschreibung
    detailed_text = scrolledtext.ScrolledText(detailed_window, wrap=tk.WORD, height=10, width=40)
    detailed_text.pack()
    detailed_text.insert(tk.END, detailed_description)
    detailed_text.config(state=tk.DISABLED)

# Funktion für die Statistik und Plot
def zeige_statistik_plot():
    # Statistik erstellen
    buchstaben_statistik = {}
    for abkuerzung in abkuerzungen_beschreibungen.keys():
        erstes_buchstabe = abkuerzung[0].upper()
        buchstaben_statistik[erstes_buchstabe] = buchstaben_statistik.get(erstes_buchstabe, 0) + 1

    # Plot erstellen
    plt.bar(buchstaben_statistik.keys(), buchstaben_statistik.values())
    plt.xlabel('Buchstabe')
    plt.ylabel('Anzahl der Abkürzungen')
    plt.title('Statistik der Abkürzungen pro Buchstabe')

    # Canvas für den Plot in der GUI einbetten
    figure_canvas = FigureCanvasTkAgg(plt.gcf(), master=statistik_tab)
    figure_canvas.draw()
    figure_canvas.get_tk_widget().pack(side=tk.TOP, fill=tk.BOTH, expand=1)

    # Warnung anzeigen, falls keine Daten für den Plot vorhanden sind
    if not buchstaben_statistik:
        messagebox.showwarning("Warnung", "Keine Daten für den Plot vorhanden.")

# Autocomplete-Funktion für das Eingabefeld
def autocomplete(entry, lista):
    current_text = entry.get().lower()
    filtered_list = [item for item in lista if current_text in item.lower()]
    entry['values'] = filtered_list

# Funktion zum Laden der Datei in das Wörterbuch beim Start
def lade_woerterbuch(dateipfad):
    return datei_in_woerterbuch_laden(dateipfad)

# GUI erstellen
root = tk.Tk()
root.title("AbbSearcher")

# Tabs hinzufügen
notebook = ttk.Notebook(root)
notebook.pack(fill=tk.BOTH, expand=True)

# Tab für die Suche
suche_tab = ttk.Frame(notebook)
notebook.add(suche_tab, text='Search')

eingabe_label = tk.Label(suche_tab, text="Enter an abbreviation:")
eingabe_label.pack()

eingabe_abkuerzung_var = tk.StringVar()
eingabe_abkuerzung = ttk.Combobox(suche_tab, textvariable=eingabe_abkuerzung_var)
eingabe_abkuerzung.pack()

# Datei in Wörterbuch laden (nur einmal)
dateipfad = "AbbList.xlsx"
abkuerzungen_beschreibungen = lade_woerterbuch(dateipfad)

eingabe_abkuerzung['values'] = list(abkuerzungen_beschreibungen.keys())
eingabe_abkuerzung.bind('<KeyRelease>', lambda event: autocomplete(eingabe_abkuerzung, list(abkuerzungen_beschreibungen.keys())))
eingabe_abkuerzung.bind("<Return>", suche_beschreibung)

suchen_button = tk.Button(suche_tab, text="Search", command=suche_beschreibung)
suchen_button.pack()

beschreibungen_text = scrolledtext.ScrolledText(suche_tab, wrap=tk.WORD, height=10, width=40)
beschreibungen_text.pack()
beschreibungen_text.config(state=tk.DISABLED)

# Tab für die Statistik
statistik_tab = ttk.Frame(notebook)
notebook.add(statistik_tab, text='Statistics & Plot')

statistik_button = tk.Button(statistik_tab, text="Show Statistics & Plot", command=zeige_statistik_plot)
statistik_button.pack()

root.mainloop()
